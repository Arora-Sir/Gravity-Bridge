import os
import io
import json
import uuid
import datetime
import requests
import pypdf
import zipfile
from bs4 import BeautifulSoup
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, FileResponse
from pydantic import BaseModel
from typing import List, Optional

app = FastAPI(title="GravityBridge Server 2.0", description="Agentic Local Backend with Tool Call capabilities")

# --- Config from .env ---
def _load_env_value(key, default=""):
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(env_path):
        return default
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line.startswith("#") or "=" not in line:
                continue
            k, _, v = line.partition("=")
            if k.strip() == key:
                val = v.strip()
                if len(val) >= 2:
                    if (val[0] == '"' and val[-1] == '"') or \
                       (val[0] == "'" and val[-1] == "'"):
                        val = val[1:-1]
                if val and not val.startswith("CHANGE_ME") and not val.startswith("C:\\your"):
                    return val
    return default

USER_HOME_PATH = _load_env_value("USER_HOME_PATH") or os.path.expanduser("~")
USER_DISPLAY_NAME = _load_env_value("USER_DISPLAY_NAME") or os.environ.get("USERNAME", "User")

# CORS setup to allow mobile browser access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

STORAGE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "storage")
UPLOAD_DIR = os.path.join(STORAGE_DIR, "uploaded_files")
EXTRACT_DIR = os.path.join(STORAGE_DIR, "extracted")
os.makedirs(STORAGE_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EXTRACT_DIR, exist_ok=True)

# Helper function to get Gemini API Key
def get_api_key(client_key: str = None) -> str:
    if client_key and client_key.strip():
        return client_key.strip()
    
    # 1. Check .env file
    env_file_key = _load_env_value("GEMINI_API_KEY")
    if env_file_key and env_file_key != "CHANGE_ME":
        return env_file_key

    # 2. Check environment variable
    env_key = os.environ.get("GEMINI_API_KEY")
    if env_key:
        return env_key

    # 3. Fallback to config.json
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, "r") as f:
                cfg = json.load(f)
                val = cfg.get("gemini_api_key", "").strip()
                if val and val != "CHANGE_ME":
                    return val
        except:
            pass

    raise HTTPException(
        status_code=400, 
        detail="Gemini API Key not found. Please set GEMINI_API_KEY in your .env file on your laptop."
    )

# Path resolver to guarantee Windows path parsing safety (relative to user home or absolute)
def resolve_path(path_str: str) -> str:
    path_str = path_str.strip()
    path_str = path_str.replace("/", os.sep)
    
    if os.path.splitdrive(path_str)[0] or path_str.startswith("\\\\"):
        return path_str
        
    return os.path.abspath(os.path.join(USER_HOME_PATH, path_str))

# --- Local Tools for Gemini Agent ---

def list_directory(path: str) -> dict:
    target = resolve_path(path)
    if not os.path.exists(target):
        return {"status": "failed", "error": f"Path not found: {path}"}
    if not os.path.isdir(target):
        return {"status": "failed", "error": f"Path is a file, not a directory: {path}"}
        
    try:
        items = []
        for name in os.listdir(target):
            if name.startswith(".") or name.startswith("$"):
                continue
            full_path = os.path.join(target, name)
            is_dir = os.path.isdir(full_path)
            size = os.path.getsize(full_path) if not is_dir else 0
            items.append({
                "name": name,
                "type": "directory" if is_dir else "file",
                "size_bytes": size
            })
        return {"status": "success", "resolved_path": target, "contents": items}
    except Exception as e:
        return {"status": "failed", "error": str(e)}

def unzip_file(path: str) -> dict:
    target = resolve_path(path)
    if not os.path.exists(target):
        return {"status": "failed", "error": f"Zip file not found: {path}"}
    if not zipfile.is_zipfile(target):
        return {"status": "failed", "error": f"Not a valid ZIP file: {path}"}
        
    temp_dir_name = f"extracted_{uuid.uuid4().hex[:8]}"
    extract_to = os.path.join(EXTRACT_DIR, temp_dir_name)
    os.makedirs(extract_to, exist_ok=True)
    
    try:
        with zipfile.ZipFile(target, 'r') as z:
            z.extractall(extract_to)
            
        extracted_files = []
        for root, _, filenames in os.walk(extract_to):
            for fname in filenames:
                full_path = os.path.join(root, fname)
                rel = os.path.relpath(full_path, extract_to)
                extracted_files.append(rel)
                
        return {
            "status": "success",
            "message": "Successfully unzipped archive.",
            "extracted_directory": extract_to,
            "files": extracted_files
        }
    except Exception as e:
        return {"status": "failed", "error": str(e)}

def read_file_content(path: str) -> dict:
    target = resolve_path(path)
    if not os.path.exists(target):
        return {"status": "failed", "error": f"File not found: {path}"}
    if os.path.isdir(target):
        return {"status": "failed", "error": f"Path is a directory: {path}"}
        
    ext = os.path.splitext(target)[1].lower()
    
    try:
        # Code & Text Files
        if ext in [".txt", ".md", ".py", ".java", ".cpp", ".c", ".h", ".cs", ".json", ".html", ".js", ".css", ".xml", ".yaml", ".yml", ".ini", ".bat", ".ps1"]:
            with open(target, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read(250000) # Read up to 250k chars
            return {"status": "success", "content": content}
            
        # PDF parsing
        elif ext == ".pdf":
            reader = pypdf.PdfReader(target)
            text = []
            for idx, page in enumerate(reader.pages):
                page_text = page.extract_text()
                text.append(f"--- PAGE {idx+1} ---\n{page_text or '[No text extracted]'}\n")
            return {"status": "success", "content": "".join(text)}
            
        # CSV parsing
        elif ext == ".csv":
            import csv
            rows = []
            with open(target, "r", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f)
                for idx, row in enumerate(reader):
                    if idx > 300: # Limit rows to prevent payload bloating
                        rows.append(["... [truncated due to size] ..."])
                        break
                    rows.append(row)
            return {"status": "success", "content": json.dumps(rows)}
            
        # Excel parsing
        elif ext in [".xlsx", ".xls"]:
            import openpyxl
            wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
            output = []
            for sheet in wb.sheetnames[:5]: # First 5 sheets only
                output.append(f"--- SHEET: {sheet} ---")
                ws = wb[sheet]
                for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if r_idx > 100:
                        output.append("[... sheet truncated ...]")
                        break
                    output.append(", ".join([str(val) if val is not None else "" for val in row]))
            return {"status": "success", "content": "\n".join(output)}
            
        # Image Fallback
        elif ext in [".png", ".jpg", ".jpeg"]:
            return {"status": "success", "message": f"[Image file: {os.path.basename(target)}]. Dimensions / metadata can be inspected, OCR skipped.", "size_bytes": os.path.getsize(target)}
            
        # Default Text Fallback
        else:
            with open(target, "r", encoding="utf-8", errors="ignore") as f:
                content = f.read(50000)
            return {"status": "success", "content": content, "warning": "Unrecognized extension, read as plain text."}
            
    except Exception as e:
        return {"status": "failed", "error": str(e)}

# Execute local tool based on tool call parameters
def execute_local_tool(name: str, args: dict) -> dict:
    if name == "list_directory":
        return list_directory(args.get("path", ""))
    elif name == "unzip_file":
        return unzip_file(args.get("path", ""))
    elif name == "read_file_content":
        return read_file_content(args.get("path", ""))
    else:
        return {"status": "failed", "error": f"Tool '{name}' not found."}

# Deep recursive tool execution loop with Gemini
def run_gemini_agent(messages: list, api_key: str, model_name: str = "gemini-3.5-flash", system_instruction: str = "", depth: int = 0) -> str:
    if depth > 10:
        return "Error: Agent reached maximum recursion depth (10). Canceled execution."
        
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={api_key}"
    headers = {"Content-Type": "application/json"}
    
    # Declarations of tools exposed to the model
    tools = [{
        "functionDeclarations": [
            {
                "name": "list_directory",
                "description": "Lists the files and subfolders in a folder. Path can be absolute (e.g. 'D:\\Study') or relative to the user's home folder (e.g. 'Downloads' or 'Downloads/my_subfolder').",
                "parameters": {
                    "type": "OBJECT",
                    "properties": {
                        "path": {
                            "type": "STRING",
                            "description": "The path to list."
                        }
                    },
                    "required": ["path"]
                }
            },
            {
                "name": "read_file_content",
                "description": "Extracts text content from any supported file type (PDF, CSV, Excel, TXT, MD, Python, Java, C++, C#, JSON, HTML). Path can be absolute or relative.",
                "parameters": {
                    "type": "OBJECT",
                    "properties": {
                        "path": {
                            "type": "STRING",
                            "description": "The path of the file to read."
                        }
                    },
                    "required": ["path"]
                }
            },
            {
                "name": "unzip_file",
                "description": "Extracts a ZIP archive into a temporary folder on the laptop, and returns the list of extracted files.",
                "parameters": {
                    "type": "OBJECT",
                    "properties": {
                        "path": {
                            "type": "STRING",
                            "description": "The path of the ZIP file to extract."
                        }
                    },
                    "required": ["path"]
                }
            }
        ]
    }]
    
    payload = {
        "contents": messages,
        "tools": tools,
        "generationConfig": {
            "temperature": 0.2
        }
    }
    
    if system_instruction:
        payload["systemInstruction"] = {
            "parts": [{"text": system_instruction}]
        }
        
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=120)
        response.raise_for_status()
        data = response.json()
        
        candidates = data.get("candidates", [])
        if not candidates:
            return "Error: No response generated by Gemini."
            
        content = candidates[0].get("content", {})
        parts = content.get("parts", [])
        
        # Check for function/tool call request
        function_calls = [p.get("functionCall") for p in parts if "functionCall" in p]
        
        if function_calls:
            messages.append({
                "role": "model",
                "parts": parts
            })
            
            function_responses = []
            for call in function_calls:
                t_name = call.get("name")
                t_args = call.get("args", {})
                
                tool_res = execute_local_tool(t_name, t_args)
                
                func_response_part = {
                    "functionResponse": {
                        "name": t_name,
                        "response": tool_res
                    }
                }
                function_responses.append(func_response_part)
                
            messages.append({
                "role": "function",
                "parts": function_responses
            })
            
            return run_gemini_agent(messages, api_key, model_name, system_instruction, depth + 1)
        else:
            return parts[0].get("text", "Error: Empty response content.")
            
    except requests.exceptions.HTTPError as he:
        try:
            err_msg = response.json().get("error", {}).get("message", str(he))
        except:
            err_msg = str(he)
        return f"Gemini API Error: {err_msg}"
    except Exception as e:
        return f"Execution Error: {str(e)}"

# Save analysis logs for history
def save_chat_record(user_prompt: str, result: str, uploaded_files: list):
    analysis_id = str(uuid.uuid4())
    timestamp = datetime.datetime.now().isoformat()
    filename = f"chat_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}_{analysis_id[:8]}.json"
    
    record = {
        "id": analysis_id,
        "timestamp": timestamp,
        "source_name": uploaded_files[0] if uploaded_files else "Direct Chat Query",
        "source_type": "chat",
        "prompt": user_prompt,
        "result": result,
        "uploaded_files": uploaded_files
    }
    
    record_path = os.path.join(STORAGE_DIR, filename)
    with open(record_path, "w", encoding="utf-8") as f:
        json.dump(record, f, indent=4)
    return record

# --- API Endpoints ---

@app.post("/api/upload")
async def upload_files(files: List[UploadFile] = File(...)):
    uploaded = []
    errors = []
    
    for file in files:
        try:
            safe_name = os.path.basename(file.filename)
            file_path = os.path.join(UPLOAD_DIR, safe_name)
            
            contents = await file.read()
            with open(file_path, "wb") as f:
                f.write(contents)
                
            uploaded.append({
                "filename": safe_name,
                "local_path": file_path,
                "size_bytes": len(contents)
            })
        except Exception as e:
            errors.append({"filename": file.filename, "error": str(e)})
            
    return {"uploaded": uploaded, "errors": errors}

class MessagePart(BaseModel):
    text: Optional[str] = None

class MessageTurn(BaseModel):
    role: str
    parts: List[MessagePart]

class ChatRequest(BaseModel):
    history: List[MessageTurn]
    model: str = "gemini-3.5-flash"

@app.post("/api/chat")
async def chat_interaction(
    req: ChatRequest,
    x_api_key: str = Header(None, alias="X-Gemini-API-Key")
):
    api_key = get_api_key(x_api_key)
    
    # 1. Fetch current list of uploaded files to expose to Gemini
    uploaded_files_list = []
    if os.path.exists(UPLOAD_DIR):
        for f in os.listdir(UPLOAD_DIR):
            f_path = os.path.join(UPLOAD_DIR, f)
            if os.path.isfile(f_path):
                uploaded_files_list.append(f"- Name: {f}\n  Path: {f_path}\n  Size: {os.path.getsize(f_path)} bytes")
                
    files_context = ""
    if uploaded_files_list:
        files_context = "The user has uploaded these files from their phone onto the laptop:\n" + "\n".join(uploaded_files_list) + "\n\n"
    else:
        files_context = "No files uploaded in this folder yet.\n\n"

    # 2. Build system instruction with laptop agent metadata
    system_instruction = (
        f"You are Antigravity Portal, a powerful AI agent running locally on {USER_DISPLAY_NAME}'s laptop (Windows).\n"
        "You are communicating with the user's mobile device via the GravityBridge web interface.\n"
        "You have local tools to read files, unzip archives, and list directories on the laptop.\n"
        f"By default, resolve all relative paths relative to the user's home folder (like 'Downloads', 'Desktop', etc.).\n"
        f"{files_context}"
        "CRITICAL RULES:\n"
        "1. If the user mentions folders or files (e.g. 'Downloads folder', 'subfolder', 'darq.zip'), USE YOUR TOOLS to search, list, unzip, or read them.\n"
        "2. Do not hallucinate file contents. If you need to read a file to answer, call read_file_content first.\n"
        "3. Provide extremely helpful, clear summaries. Use markdown tags correctly.\n"
        "4. If you unzip a file, list the contents, then ask the user if they want you to read any specific code or document from the archive."
    )

    # 3. Map request history to Gemini contents structure
    gemini_history = []
    for turn in req.history:
        gemini_parts = []
        for p in turn.parts:
            if p.text:
                gemini_parts.append({"text": p.text})
        if gemini_parts:
            gemini_history.append({
                "role": turn.role,
                "parts": gemini_parts
            })
            
    if not gemini_history:
        raise HTTPException(status_code=400, detail="Empty conversation history.")

    # Get the last user message to log/track
    last_user_message = ""
    for turn in reversed(gemini_history):
        if turn["role"] == "user":
            last_user_message = "".join([p.get("text", "") for p in turn["parts"]])
            break

    # 4. Execute the agent loop
    result = run_gemini_agent(
        messages=gemini_history,
        api_key=api_key,
        model_name=req.model,
        system_instruction=system_instruction
    )

    # 5. Log record in local history files
    recent_uploads = [f for f in os.listdir(UPLOAD_DIR)]
    save_chat_record(last_user_message, result, recent_uploads)
    
    return {"result": result}

@app.get("/api/history")
async def get_history():
    records = []
    if not os.path.exists(STORAGE_DIR):
        return records
        
    for filename in sorted(os.listdir(STORAGE_DIR), reverse=True):
        if filename.startswith("chat_") and filename.endswith(".json"):
            try:
                with open(os.path.join(STORAGE_DIR, filename), "r", encoding="utf-8") as f:
                    records.append(json.load(f))
            except:
                pass
    return records

@app.delete("/api/history/{record_id}")
async def delete_history(record_id: str):
    if not os.path.exists(STORAGE_DIR):
        raise HTTPException(status_code=404, detail="Storage directory not found.")
        
    for filename in os.listdir(STORAGE_DIR):
        if filename.endswith(".json"):
            file_path = os.path.join(STORAGE_DIR, filename)
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    record = json.load(f)
                    if record.get("id") == record_id:
                        os.remove(file_path)
                        return {"status": "deleted", "id": record_id}
            except:
                pass
    raise HTTPException(status_code=404, detail="Record not found.")

# Mount the static files
static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
os.makedirs(static_dir, exist_ok=True)

@app.get("/")
async def read_index():
    index_path = os.path.join(static_dir, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return JSONResponse(content={"message": "GravityBridge server running. Static frontend not yet loaded."})

app.mount("/static", StaticFiles(directory=static_dir), name="static")
